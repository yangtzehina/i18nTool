from openpyxl import load_workbook, Workbook
import lxml.etree as etree
import pandas as pd
from openpyxl.styles import PatternFill
import os
from io import BytesIO  # Add this import

"""
核心功能模块: 包含所有XML和Excel处理功能
"""

class XMLProcessor:
    """XML处理相关的功能"""
    
    @staticmethod
    def create_entry(key, value):
        """创建一个单独的entry元素"""
        entry = etree.Element("entry")
        key_elem = etree.SubElement(entry, "KEY")
        key_elem.text = str(key)
        value_elem = etree.SubElement(entry, "VALUE1")
        value_elem.text = str(value)
        return entry
    
    @staticmethod
    def save_xml_file(root, output_path):
        """保存XML文件并添加XML声明"""
        etree.ElementTree(root).write(output_path, pretty_print=True, encoding="utf-8")
        
        # 添加XML声明
        with open(output_path, 'r', encoding='utf-8') as file:
            xml_content = file.read()

        if not xml_content.startswith('<?xml'):
            xml_content = '<?xml version="1.0" encoding="UTF-8" standalone="no" ?>\n' + xml_content
            with open(output_path, 'w', encoding='utf-8') as file:
                file.write(xml_content)
    
    @staticmethod
    def excel_to_xml(input_path, output_path):
        """将Excel文件转换为XML格式"""
        try:
            # 读取Excel文件
            df = pd.read_excel(input_path)
            # 获取列名作为语言标识
            languages = df.columns[1:]  # 跳过第一列（键名列）
            key_column = df.iloc[:, 1]  # 获取第一列作为键名
            
            # 为每种语言生成对应的XML文件
            for col_idx, language in enumerate(languages, start=1):
                # 创建输出文件名
                output_file = os.path.join(output_path, f'UILanguage_{language}.xml')
    
                # 获取当前语言的值
                value_column_data = df.iloc[:, col_idx]
    
                # 创建XML内容
                xml_content = '<?xml version="1.0" encoding="utf-8"?>\n<resources>\n'
    
                # 遍历数据并添加到XML中
                for index, key in key_column.items():
                    value = value_column_data[index]
                    # 确保值不是NaN且键不为空
                    if pd.notna(value) and pd.notna(key) and str(key).strip():
                        # 将键名作为属性名
                        xml_content += f'  <string {str(key)}>{str(value)}</string>\n'
    
                xml_content += '</resources>'
    
                # 保存XML文件
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write(xml_content)
    
                print(f"已生成语言文件: {output_file}")

            print("所有语言文件生成完成！")
            return True
        except Exception as e:
            print(f"Excel转XML失败: {str(e)}")
            raise
    
    @staticmethod
    def excel_to_xml_game(input_path, output_path):
        """
        将Excel文件转换为XML格式，每个条目包含 KEY, ID, Value1, Value2, Tag.
        输出格式:
        <root xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
          <data>
            <Template>
              <entry>
                <KEY>key_value_from_excel</KEY>
                <ID>id_value_from_excel</ID>
                <Value1>value1_from_excel</Value1>
                <Value2>value2_from_excel</Value2>
                <Tag>tag_value_from_excel</Tag>
              </entry>
              ...
            </Template>
          </data>
        </root>
        """
        try:
            df = pd.read_excel(input_path)

            required_columns = ["KEY", "ID", "Value1", "Value2", "Tag"]
            
            # Validate that the required columns exist
            for col_name in required_columns:
                if col_name not in df.columns:
                    raise ValueError(f"Excel file missing a required column: {col_name}")

            # Create the root element for the XML with xsi namespace
            NSMAP = {'xsi': "http://www.w3.org/2001/XMLSchema-instance"}
            root_xml_element = etree.Element("root", nsmap=NSMAP)
            
            data_element = etree.SubElement(root_xml_element, "data")
            template_element = etree.SubElement(data_element, "Template")

            # Iterate through each row in the DataFrame
            for index, row_data in df.iterrows():
                entry_element = etree.SubElement(template_element, "entry") # Add entry to Template
                
                key_val = row_data.get("KEY", "")
                id_val = row_data.get("ID", "")
                value1_val = row_data.get("Value1", "")
                value2_val = row_data.get("Value2", "")
                tag_val = row_data.get("Tag", "")

                # Create sub-elements for each required column
                key_sub_element = etree.SubElement(entry_element, "KEY")
                key_sub_element.text = str(key_val) if pd.notna(key_val) else ""
                
                id_sub_element = etree.SubElement(entry_element, "ID")
                id_sub_element.text = str(id_val) if pd.notna(id_val) else ""
                
                value1_sub_element = etree.SubElement(entry_element, "Value1")
                value1_sub_element.text = str(value1_val) if pd.notna(value1_val) else ""
                
                value2_sub_element = etree.SubElement(entry_element, "Value2")
                value2_sub_element.text = str(value2_val) if pd.notna(value2_val) else ""
                
                tag_sub_element = etree.SubElement(entry_element, "Tag")
                tag_sub_element.text = str(tag_val) if pd.notna(tag_val) else ""

            # Create an ElementTree object
            # tree = etree.ElementTree(root_xml_element) # Not strictly needed if passing root_xml_element to save_xml_file
            
            # Define the output file name.
            if os.path.isdir(output_path):
                output_file_path = os.path.join(output_path, 'strings-zh_tc.xml')
            else:
                # Ensure the directory for the output file exists if a full path is given
                output_dir = os.path.dirname(output_path)
                if output_dir: # Check if output_dir is not an empty string
                    os.makedirs(output_dir, exist_ok=True)
                output_file_path = output_path

            # Save XML file using the existing save_xml_file method for pretty printing and declaration
            XMLProcessor.save_xml_file(root_xml_element, output_file_path)
    
            print(f"游戏 已生成XML文件: {output_file_path}")
            return True
        except Exception as e:
            print(f"Excel转XML (game format) 失败: {str(e)}")
            raise
    

    @staticmethod
    def xml_to_excel(input_file, output_file):
        """将XML文件转换为Excel格式"""
        try:
            # 解析XML文件
            tree = etree.parse(input_file)
            
            # 查找所有entry元素
            xpath_expression = "/root/data/LanguageStringConvertor/entry"
            items = tree.xpath(xpath_expression)

            # 创建新的Excel文件
            wb = Workbook()
            ws = wb.active
            ws.append(["KEY", "VALUE1"])  # 添加标题行

            # 提取XML数据并添加到Excel
            for item in items:
                id_value = item.find("KEY").text if item.find("KEY") is not None else ""
                name_value = item.find("VALUE1").text if item.find("VALUE1") is not None else ""
                ws.append([id_value, name_value])
            
            # 保存Excel文件
            wb.save(output_file)
            return True
        except Exception as e:
            print(f"XML转Excel失败: {str(e)}")
            raise


class ExcelProcessor:
    """Excel处理相关的功能"""
    
    @staticmethod
    def clean_sheet(sheet):
        """清理Excel表格中的空行和空列"""
        # 删除空行
        rows_to_delete = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row))
        for row in reversed(rows_to_delete):
            if all(cell.value is None for cell in row):
                sheet.delete_rows(row[0].row, amount=1)

        # 删除空列
        cols_to_delete = list(sheet.iter_cols(min_col=1, max_col=sheet.max_column))
        for col in reversed(cols_to_delete):
            if all(cell.value is None for cell in col):
                sheet.delete_cols(col[0].col_idx, amount=1)
        
        return sheet
    
    @staticmethod
    def compare_excel(input_file, dist_file):
        """
        比较两个Excel表格的KEY和VALUE1字段，直接更新dist_file文件，
        并用颜色标记新增和修改的内容。
        """
        try:
            # 获取文件
            source = pd.read_excel(input_file)
            dist = pd.read_excel(dist_file)
            
            # 确保两个表格都有KEY和VALUE1列
            if 'KEY' not in source.columns or 'VALUE1' not in source.columns:
                raise ValueError("源文件缺少KEY或VALUE1列")
            if 'KEY' not in dist.columns or 'VALUE1' not in dist.columns:
                raise ValueError("目标文件缺少KEY或VALUE1列")
            
            # 创建源表的键值对字典
            source_dict = dict(zip(source['KEY'], source['VALUE1']))
            
            # 得到目标表原始列
            original_columns = dist.columns.tolist()
            
            # 创建结果表的副本
            result = dist.copy()
            
            # 记录需要标记的单元格
            cells_to_highlight = []  # [(行索引, "新增"或"修改")]
            
            # 检查修改的内容
            for idx, row in result.iterrows():
                key = row['KEY']
                if key in source_dict and row['VALUE1'] != source_dict[key]:
                    # 值不同 - 更新值
                    result.at[idx, 'VALUE1'] = source_dict[key]
                    cells_to_highlight.append((idx, "修改"))
            
            # 检查新增的内容
            existing_keys = set(result['KEY'])
            new_rows = []
            for key, value in source_dict.items():
                if key not in existing_keys:
                    # 准备新行数据
                    new_row = {col: None for col in original_columns}
                    new_row['KEY'] = key
                    new_row['VALUE1'] = value
                    new_rows.append(new_row)
            
            # 添加新行到结果
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                result = pd.concat([result, new_df], ignore_index=True)
                # 记录新增行的索引
                for i in range(len(new_rows)):
                    cells_to_highlight.append((len(dist) + i, "新增"))
            
            # 导出为Excel (覆盖原文件)
            result.to_excel(dist_file, index=False)
            
            # 打开并设置颜色
            wb = load_workbook(dist_file)
            ws = wb.active
            
            # 定义填充颜色
            new_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 绿色
            changed_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
            
            # 为标记的单元格设置颜色
            for idx, status in cells_to_highlight:
                # 调整行索引：Excel行从1开始，第1行是标题，所以+2
                excel_row = idx + 2
                
                # 为整行设置颜色
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=excel_row, column=col)
                    if status == "新增":
                        cell.fill = new_fill
                    elif status == "修改":
                        cell.fill = changed_fill
            
            # 保存结果 (覆盖原文件)
            wb.save(dist_file)
            
            # 返回修改的统计信息
            modifications = sum(1 for _, status in cells_to_highlight if status == "修改")
            new_entries = sum(1 for _, status in cells_to_highlight if status == "新增")
            
            stats = {
                "modifications": modifications,
                "new_entries": new_entries
            }
            
            return stats
        except Exception as e:
            print(f"Excel对比失败: {str(e)}")
            raise

    @staticmethod
    def compare_xml_excel(input_file, dist_file):
        """
        比较xml Excel相同的key的value值，直接更新dist_file文件，
        并用颜色标记新增和修改的内容。
        """
        try:
            # Read the XML file content as a string
            with open(input_file, 'r', encoding='utf-8') as f:
                xml_content_str = f.read()
            
            # Remove the ETX character (ASCII 03)
            cleaned_xml_content_str = xml_content_str.replace('\x03', '')
            
            # Convert the cleaned string to bytes and use BytesIO to make it file-like
            cleaned_xml_content_bytes = cleaned_xml_content_str.encode('utf-8')
            xml_file_like_object = BytesIO(cleaned_xml_content_bytes)

            # 解析XML文件 from the cleaned, file-like object
            tree = etree.parse(xml_file_like_object)
            
            # 查找所有entry元素
            xpath_expression = "//string"
            items = tree.xpath(xpath_expression)
            
            # 获取文件
            dist = pd.read_excel(dist_file)
            
            if 'KEY' not in dist.columns or 'VALUE1' not in dist.columns:
                raise ValueError("目标文件缺少KEY或VALUE1列")
            
            # 创建源表的键值对字典
            source_dict = {}
            
            # 提取XML数据并添加到字典
            for string in items:
                # 构建完整的属性字符串作为KEY
                attributes = []
                for attr_name, attr_value in string.attrib.items():
                    attributes.append(f'{attr_name}="{attr_value}"')
    
                key = " ".join(attributes)  # 组合所有属性为一个完整的字符串
                # 获取标签内容作为VALUE
                value = string.text if string.text else ""
            
                # 添加到字典
                if key:  # 只有当name属性存在时才添加
                    source_dict[key] = value

            # 打印源字典内容
            print("Source dictionary contents:")
            print(source_dict)
            # 得到目标表原始列
            original_columns = dist.columns.tolist()
            
            # 创建结果表的副本
            result = dist.copy()
            
            # 记录需要标记的单元格
            cells_to_highlight = []  # [(行索引, "新增"或"修改")]
            
            # 检查修改的内容
            for idx, row in result.iterrows():
                key = row['KEY']
                if key in source_dict and row['VALUE1'] != source_dict[key]:
                    # 值不同 - 更新值
                    result.at[idx, 'VALUE1'] = source_dict[key]
                    cells_to_highlight.append((idx, "修改"))
            
            # 检查新增的内容
            existing_keys = set(result['KEY'])
            new_rows = []
            for key, value in source_dict.items():
                if key not in existing_keys:
                    # 准备新行数据
                    new_row = {col: None for col in original_columns}
                    new_row['KEY'] = key
                    new_row['VALUE1'] = value
                    new_rows.append(new_row)
            
            # 添加新行到结果
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                result = pd.concat([result, new_df], ignore_index=True)
                # 记录新增行的索引
                for i in range(len(new_rows)):
                    cells_to_highlight.append((len(dist) + i, "新增"))
            
            # 导出为Excel (覆盖原文件)
            result.to_excel(dist_file, index=False)
            
            # 打开并设置颜色
            wb = load_workbook(dist_file)
            ws = wb.active
            
            # 定义填充颜色
            new_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # 绿色
            changed_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
            
            # 为标记的单元格设置颜色
            for idx, status in cells_to_highlight:
                # 调整行索引：Excel行从1开始，第1行是标题，所以+2
                excel_row = idx + 2
                
                # 为整行设置颜色
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=excel_row, column=col)
                    if status == "新增":
                        cell.fill = new_fill
                    elif status == "修改":
                        cell.fill = changed_fill
            
            # 保存结果 (覆盖原文件)
            wb.save(dist_file)
            
            # 返回修改的统计信息
            modifications = sum(1 for _, status in cells_to_highlight if status == "修改")
            new_entries = sum(1 for _, status in cells_to_highlight if status == "新增")
            
            stats = {
                "modifications": modifications,
                "new_entries": new_entries
            }
            
            return stats
        except Exception as e:
            print(f"Excel对比失败: {str(e)}")
            raise

# 公共API函数，供其他模块调用
def convert_xml_to_excel(input_file, dist_file):
    """比较和更新xml2Excel文件"""
    stats = ExcelProcessor.compare_xml_excel(input_file, dist_file)
    print(f"文件已更新: {dist_file}")
    print(f"已修改 {stats['modifications']} 个条目，新增 {stats['new_entries']} 个条目")
    return stats

def convert_excel_to_xml(input_file, output_file):
    """将Excel文件转换为XML文件"""
    return XMLProcessor.excel_to_xml(input_file, output_file)

def convert_excel_to_xml_game(input_file, output_file):
    """将Excel文件转换为游戏特定格式的XML文件"""
    return XMLProcessor.excel_to_xml_game(input_file, output_file)

def compare_language_excel(input_file, dist_file):
    """比较和更新Excel文件"""
    stats = ExcelProcessor.compare_excel(input_file, dist_file)
    print(f"文件已更新: {dist_file}")
    print(f"已修改 {stats['modifications']} 个条目，新增 {stats['new_entries']} 个条目")
    return stats

